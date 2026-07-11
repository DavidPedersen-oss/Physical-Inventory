import openpyxl, json, datetime, re

SRC="survey.xlsx"
wb=openpyxl.load_workbook(SRC, read_only=True, data_only=True)

def s(x):
    if x is None: return ''
    if isinstance(x,(datetime.datetime,datetime.date)): return x.isoformat()[:10]
    if isinstance(x,float) and x==int(x): return str(int(x))
    return str(x).strip()

# category + archive per sheet, and column-name aliases
SHEETS={
 'NEW SURVEY APP':('NEW',0),'Bookstore':('BK',0),' CMP-INV ':('CI',0),
 'CMP-NON-INV':('CNI',0),'TRANSFERS':('TR',0),' FDN':('FDN',0),
 'CMP - INV (ARCHIVE)':('CI',1),'CMP - NON-INV (ARCHIVE)':('CNI',1),
 'TRANSFERS (ARCHIVES)':('TR',1),'FDN (ARCHIVE)':('FDN',1),
}
CATNAME={'NEW':'New Survey App','BK':'Bookstore E-Waste','CI':'Campus Inventory',
 'CNI':'Campus Non-Inventory','TR':'Transfer','FDN':'Foundation'}

def find(hdr,*keys):
    for i,h in enumerate(hdr):
        hl=h.lower().replace('\n',' ')
        for k in keys:
            if k in hl: return i
    return None

current=[]; archive=[]; KEY=0
for name in wb.sheetnames:
    cat,ar=SHEETS[name]; ws=wb[name]
    rows=list(ws.iter_rows(min_row=2,max_row=2,values_only=True))
    hdr=[s(c) for c in rows[0]] if rows else []
    C={
     'no':find(hdr,'survey #','survey#'),'d':find(hdr,'description'),
     'aid':find(hdr,'asset id'),'tag':find(hdr,'tag'),
     'ser':find(hdr,'serial'),'yr':find(hdr,'in service','acq date','acqdate'),
     'amt':find(hdr,'original amount','amount'),
     'da':find(hdr,'date assigned'),'dc':find(hdr,'date completed'),
     'lb':find(hdr,'logged'),'cb':find(hdr,'closed'),
     'cc':find(hdr,'disposal code','disposal\ncode'),'cond':find(hdr,'disposal condition','condition'),
     'act':find(hdr,'disposal action','action'),'nt':find(hdr,'notes'),
     'fy':find(hdr,'fy sort','fy'),
    }
    # dept: transfers have two dept ids (from/to). generic dept id/name
    dept_idx=[i for i,h in enumerate(hdr) if 'dept id' in h.lower().replace('\n',' ') or 'dept #' in h.lower().replace('\n',' ') or 'dept#' in h.lower().replace('\n',' ')]
    dname_idx=[i for i,h in enumerate(hdr) if 'dept name' in h.lower().replace('\n',' ')]
    is_tr=(cat=='TR')
    def g(r,i):
        return s(r[i]) if (i is not None and i<len(r)) else ''
    for r in ws.iter_rows(min_row=3,values_only=True):
        if all(c is None for c in r): continue
        desc=g(r,C['d']); no=g(r,C['no'])
        tag=g(r,C['tag']); ser=g(r,C['ser']); aid=g(r,C['aid'])
        if not (desc or no or tag or ser): continue
        rec={'k':KEY,'cat':cat,'ar':ar}
        if no: rec['no']=no
        if desc: rec['d']=desc
        if aid: rec['aid']=aid
        if tag: rec['tag']=tag
        if ser: rec['ser']=ser
        for f in ('yr','amt','da','dc','lb','cb','cc','cond','act','nt','fy'):
            v=g(r,C[f])
            if v: rec[f]=v
        # dept handling
        if is_tr and len(dept_idx)>=2:
            fd=g(r,dept_idx[0]); td=g(r,dept_idx[1])
            fdn=g(r,dname_idx[0]) if len(dname_idx)>=1 else ''
            tdn=g(r,dname_idx[1]) if len(dname_idx)>=2 else ''
            if fd: rec['fd']=fd
            if td: rec['td']=td
            if fdn: rec['fdn']=fdn
            if tdn: rec['tdn']=tdn
        else:
            dp=g(r,dept_idx[0]) if dept_idx else ''
            dn=g(r,dname_idx[0]) if dname_idx else ''
            if dp: rec['dp']=dp
            if dn: rec['dn']=dn
        # status: open if current sheet and no completion date
        rec['st']= 'Open' if (not ar and not g(r,C['dc'])) else 'Done'
        (archive if ar else current).append(rec)
        KEY+=1

print("current recs:",len(current)," archive recs:",len(archive))

# ---------- matching to assets ----------
def digits(x): 
    d=re.sub(r'\D','',x or ''); return d.lstrip('0')
def alnum(x): return re.sub(r'[^A-Z0-9]','',(x or '').upper())
def idn(x):
    x=(x or '').strip(); 
    return x.lstrip('0') or x

assets=json.load(open('PI/seed/assets.json'))+json.load(open('PI/seed/disposed.json'))
byId={}; byTag={}; bySer={}
for a in assets:
    i=idn(str(a.get('id','')))
    if i and i not in byId: byId[i]=(a['id'],a.get('lc','A'))
    t=digits(str(a.get('tag','')))
    if len(t)>=3 and t not in byTag: byTag[t]=(a['id'],a.get('lc','A'))
    se=alnum(str(a.get('serial','')))
    if len(se)>=5 and se not in bySer: bySer[se]=(a['id'],a.get('lc','A'))

def match(rec):
    i=idn(rec.get('aid',''))
    if i and i in byId: return byId[i]+('id',)
    t=digits(rec.get('tag',''))
    if len(t)>=3 and t in byTag: return byTag[t]+('tag',)
    se=alnum(rec.get('ser',''))
    if len(se)>=5 and se in bySer: return bySer[se]+('serial',)
    return None

stats={'id':0,'tag':0,'serial':0}
for pool in (current,archive):
    for rec in pool:
        mm=match(rec)
        if mm:
            rec['m']=mm[0]; rec['ml']=mm[1]; rec['mt']=mm[2]; stats[mm[2]]+=1

matched_cur=sum(1 for r in current if 'm' in r)
matched_arc=sum(1 for r in archive if 'm' in r)
print("matched current:",matched_cur,"/",len(current))
print("matched archive:",matched_arc,"/",len(archive))
print("by type:",stats)

json.dump(current,open('PI/seed/surveys.json','w'),separators=(',',':'))
json.dump(archive,open('PI/seed/surveys_archive.json','w'),separators=(',',':'))
import os
print("surveys.json      %.2f MB"%(os.path.getsize('PI/seed/surveys.json')/1e6))
print("surveys_archive   %.2f MB"%(os.path.getsize('PI/seed/surveys_archive.json')/1e6))

# survey-# grouping preview
from collections import defaultdict
grp=defaultdict(int)
for r in current:
    grp[(r['cat'],r.get('no','(none)'))]+=1
print("distinct current survey groups:",len(grp))
